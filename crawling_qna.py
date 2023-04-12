# required to install selenium, bs4 
from selenium import webdriver
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


import openpyxl
from openpyxl.styles import PatternFill, Color
from openpyxl import Workbook
from random import *


# Driver Setting
profile = webdriver.FirefoxProfile()

profile.set_preference('general.useragent.override', 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:65.0) Gecko/20100101 Firefox/65.0')
profile.set_preference("network.proxy.type", 1)
profile.set_preference("network.proxy.socks", "127.0.0.1")
profile.set_preference("network.proxy.socks_port", 9050)

path = "./driver/geckodriver"
driver = webdriver.Firefox(firefox_profile=profile, executable_path=path)


# Crawling each q&a => Write on excel
date = str(datetime.now()).replace('.', '_')
date = date.replace(' ', '_')
filename = './result/' +  date + "_crawling_result.xlsx"
## excel workbook setting
wb = Workbook()
sheet = wb.active
sheet.append(['제목', '질문', '답변'])
for j in range(1, 4):
    sheet.cell(row=1, column=j).fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
## read url list in file
f = open("./result/url_list_2023-04-09_00:21:59_710019.txt", 'r')
page_url = f.readlines()
f.close()
## crawling each url page
for i in page_url[4056:5000]:
    time.sleep(uniform(1.0, 2.0))
    driver.get(i)
    # Question list
    # alert창 처리
    '''
    try:
        result = driver.switch_to.alert()
        continue
    except:
        pass
    '''
    #title = driver.find_element_by_class_name('title').text
    try:
        title = driver.find_element(By.CLASS_NAME,'title').text
    except:
        title = "@Exception@"
    try:
        #question_txt = driver.find_element_by_class_name('c-heading__content').text
        question_txt = driver.find_element(By.CLASS_NAME,'c-heading__content').text
    except:
        question_txt = "@Exception@"

    # Answer list
    #answer_list = driver.find_elements_by_class_name("se-main-container")
    answer_list = driver.find_elements(By.CLASS_NAME, "se-main-container")
    for n, answer in enumerate(answer_list):
        t = ""
        #texts = answer.find_elements_by_tag_name('span')
        try:
            texts = answer.find_elements(By.TAG_NAME, 'span')
        except:
            t="@Excepiton"
        for i in texts:
            t += i.text
        try:
            sheet.append([title, question_txt, t])
        except:
            print("Exception for writing to excel")

    wb.save(filename)